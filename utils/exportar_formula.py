import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

def exportar_hoja_trabajo_excel(df: pd.DataFrame, nombre_formula: str, codigo: str = "", fecha: str = "", logo_path: str = "logo.png") -> BytesIO:
    wb = Workbook()
    ws = wb.active

    # Insertar logo en A1 (3x5 cm aprox)
    try:
        logo = XLImage(logo_path)
        logo.width = 150  # ~5cm
        logo.height = 90  # ~3cm
        ws.add_image(logo, "A1")
    except Exception as e:
        print(f"Error cargando logo: {e}")

    # Título en C3
    titulo = nombre_formula
    if codigo:
        titulo += f" - MUESTRA: {codigo}"
    ws.cell(row=3, column=3, value=titulo)

    # Fecha en H3 (columna 8)
    if fecha:
        ws.cell(row=3, column=8, value=f"FECHA: {fecha}")

    # Encabezados de tabla
    ws.cell(row=5, column=1, value="Órden de adición")
    ws.cell(row=5, column=4, value="Cantidad % peso")

    # Cargar ingredientes y porcentajes
    start_row = 6
    for idx, row in df.iterrows():
        ws.cell(row=start_row + idx, column=1, value=idx + 1)  # Órden
        ws.cell(row=start_row + idx, column=3, value=row.get("Materia Prima", ""))
        ws.cell(row=start_row + idx, column=4, value=row.get("%", 0))

    # Fila final con Densidad y pH
    end_data_row = start_row + len(df)
    ws.cell(row=end_data_row + 1, column=4, value="Densidad")
    ws.cell(row=end_data_row + 2, column=4, value="pH")

    # Ajuste de ancho de columnas
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def exportar_formula_excel(df: pd.DataFrame, nombre_formula: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fórmula"

    # Mantener solo columnas útiles
    columnas_base = ["Materia Prima", "%", "Precio €/kg"]
    columnas_tecnicas = [col for col in df.columns if col not in columnas_base and col != "id"]
    columnas_final = columnas_base + columnas_tecnicas

    # Escribir encabezados
    for col_idx, col_name in enumerate(columnas_final, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Escribir datos
    for row_idx, row in enumerate(df[columnas_final].itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Calcular rango de filas
    start_row = 2
    end_row = start_row + len(df) - 1

    # Fila en blanco + fila de totales
    total_row = end_row + 2
    ws.cell(row=total_row, column=1, value="TOTAL")

    col_pct_idx = columnas_final.index("%") + 1
    pct_col_letter = get_column_letter(col_pct_idx)
    rango_pct = f"${pct_col_letter}${start_row}:${pct_col_letter}${end_row}"

    for col_name in columnas_tecnicas:
        col_idx = columnas_final.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        formula = f"=SUMPRODUCT({rango_pct},{col_letter}{start_row}:{col_letter}{end_row})/100"
        ws.cell(row=total_row, column=col_idx, value=formula)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

